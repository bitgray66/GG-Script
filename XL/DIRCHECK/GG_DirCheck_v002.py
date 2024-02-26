import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
dt = datetime.datetime.now()

"""
Model Anim
Lo -Anim

Model LRC
Render - Proxy - Hi

TX Anim
Lo - viewport - An

TX LRC
Hi
"""

def check_f_path(row):


    path = row[13]

        
    if path:

        f_path = os.path.join("R:", path, "Final")


        if os.path.exists(f_path):


            anim_path = os.path.join(f_path, "Anim")
            lo_path = os.path.join(f_path, "Lo")
            render_path = os.path.join(f_path, "Render")
            proxy_path = os.path.join(f_path, "Proxy")
            hi_path = os.path.join(f_path, "Hi")
            
            if (os.path.exists(anim_path) or os.path.exists(lo_path)) and (os.path.exists(render_path) or os.path.exists(proxy_path) or os.path.exists(hi_path)):


                return "ALRPHP"
            elif os.path.exists(anim_path) or os.path.exists(lo_path):


                return "ALP"
            elif os.path.exists(render_path) or os.path.exists(proxy_path) or os.path.exists(hi_path):


                return "RPHP"
            else:

                return "ALRPHA"
        else:

            return "FM"
    else:

        return "PM"



def check_s_path(row):


    path = row[13]


        
    if path:

        s_path = os.path.join("R:", path, "Sourceimages")


    
        if os.path.exists(s_path):


            viewport_path = os.path.join(s_path, "Viewport")
            tx_lo_path = os.path.join(s_path, "Lo")
            tx_an_path = os.path.join(s_path, "An")
            tx_hi_path = os.path.join(s_path, "Hi")
            
            if (os.path.exists(viewport_path) or os.path.exists(tx_lo_path) or os.path.exists(tx_an_path)) and os.path.exists(tx_hi_path):


                return "VLHP"
            elif os.path.exists(viewport_path) or os.path.exists(tx_lo_path) or os.path.exists(tx_an_path):


                return "VLP"
            elif os.path.exists(tx_hi_path):


                return "HP"
            else:

                return "VLHA"
        else:

            return "SM"
    else:

        return "PM"

def process_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook[sheetName]

    msg=""
    i=2
    for row in sheet.iter_rows(min_row=2, max_row=nRow, min_col=1, max_col=14, values_only=True):

        f_result = check_f_path(row)
        s_result = check_s_path(row)

        
        match f_result:
            case "PM":
                msg = "Path mancante!"
                for a in range(1, 5):
                    sheet.cell(row=i, column=a, value=msg).fill=gray
            case "FM":
                msg = "Cartella 'Final' mancante"
                sheet.cell(row=i, column=2, value=msg).fill=red
                sheet.cell(row=i, column=3, value=msg).fill=red
            case "ALRPHA":
                msg12 = "cartella 'Anim/Lo' assente"
                msg24 = "cartella 'Render/Proxy/Hi' assente"
                sheet.cell(row=i, column=2, value=msg12).fill=orange
                sheet.cell(row=i, column=3, value=msg24).fill=orange
            case "ALRPHP":
                msg12 = "cartella 'Anim/Lo' presente"
                msg24 = "cartella 'Render/Proxy/Hi' presente"
                sheet.cell(row=i, column=2, value=msg12).fill=green
                sheet.cell(row=i, column=3, value=msg24).fill=green
            case "ALP":
                msg12 = "cartella 'Anim/Lo' presente"
                msg24 = "cartella 'Render/Proxy/Hi' assente"
                sheet.cell(row=i, column=2, value=msg12).fill=green
                sheet.cell(row=i, column=3, value=msg24).fill=orange
            case "RPHP":
                msg12 = "cartella 'Anim/Lo' assente"
                msg24 = "cartella 'Render/Proxy/Hi' presente"
                sheet.cell(row=i, column=2, value=msg12).fill=orange
                sheet.cell(row=i, column=3, value=msg24).fill=green

        

        match s_result:
            case "PM":
                msg = "Path mancante!"
                for a in range(1, 5):
                    sheet.cell(row=i, column=a, value=msg)
            case "SM":
                msg = "Cartella 'Sourceimages' mancante"
                sheet.cell(row=i, column=4, value=msg).fill=red
                sheet.cell(row=i, column=5, value=msg).fill=red
            case "VLHA":
                msg3 = "cartella 'Viewport/Lo' assente"
                msg4 = "cartella 'Hi' assente"
                sheet.cell(row=i, column=4, value=msg3).fill=orange
                sheet.cell(row=i, column=5, value=msg4).fill=orange
            case "VLHP":
                msg3 = "cartella 'Viewport/Lo/An' presente"
                msg4 = "cartella 'Hi' presente"
                sheet.cell(row=i, column=4, value=msg3).fill=green
                sheet.cell(row=i, column=5, value=msg4).fill=green
            case "VLP":
                msg3 = "cartella 'Viewport/Lo/An' presente"
                msg4 = "cartella 'Hi' assente"
                sheet.cell(row=i, column=4, value=msg3).fill=green
                sheet.cell(row=i, column=5, value=msg4).fill=orange
            case "HP":
                msg3 = "cartella 'Viewport/Lo' assente"
                msg4 = "cartella 'Hi' presente"
                sheet.cell(row=i, column=4, value=msg3).fill=orange
                sheet.cell(row=i, column=5, value=msg4).fill=green
    
        i=i+1
    workbook.save(output_path)

if __name__ == "__main__":
    
    green = PatternFill(fill_type='solid', start_color='92D050')
    orange = PatternFill(fill_type='solid', start_color='FBBC04')
    red = PatternFill(fill_type='solid', start_color='EA3916')
    gray = PatternFill(fill_type='solid', start_color='d3d3d3')
    sheetName = "SET EL PROP"
    nRow =1125
    nEp = 133
    excel_file_path = "c:/Users/luigi.marazzi/Desktop/GG/XL/DIRCHECK/ME_ASSET_LIST.xlsx"
    output_path = "c:/Users/luigi.marazzi/Desktop/GG/XL/DIRCHECK/DirCheck_"+str(nEp)+"_%s%s_%s%s%s.xlsx" % ( dt.hour, dt.minute, dt.day, dt.month, dt.year)
    process_excel(excel_file_path)

    """ Settare la largezza delle 4 col a 30  """