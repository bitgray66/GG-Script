"""  ===========================================================================
 Copyright 2023 Mobo, Inc. All rights reserved.

 Use of this software is subject to the terms of the Mobo license
 agreement provided at the time of installation or download, or which
 otherwise accompanies this software in either electronic or hard copy form.
 ===========================================================================

    Creation Date:  15 Dec 2023
    Author:
      Luigi Marazzi   

    Description:
        This procedure ....

    Input Arguments:
        The ......

    Return Value:
        .............

//////////////////////////////////////////////////////////////
  Expand/collapse procedures section
///////////////////////////////////////////////////////////// """

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import datetime
dt = datetime.datetime.now()

def compare_and_highlight(file1, sheet1, col1, file2, sheet2, col2):
    
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)


    sheet_1 = wb1[sheet1]
    sheet_2 = wb2[sheet2]


    values_col1 = [cell.value for cell in sheet_1[col1] if cell.value is not None]

    values_col2 = [cell.value for cell in sheet_2[col2] if cell.value is not None]


    common_values = set(values_col1).intersection(values_col2)


    nr =2
    for row in range(2, n_rowFtp + 1):

        cell_value = sheet_1.cell(row=row, column=2).value
        cat_value = sheet_1.cell(row=row, column=3).value


        if cat_value == "Props" or cat_value == "SetElements":

            if cell_value in common_values:
                if(sheet_1.cell(row=row, column=2).value != None):
                    sheet_1.cell(row=row, column=1).fill = green
                    sheet_1.cell(row=row, column=1, value="OK")
            else:
                if(sheet_1.cell(row=row, column=2).value != None):
                    sheet_1.cell(row=row, column=1).fill = red
                    sheet_1.cell(row=row, column=1, value="ToDo")
        
        nr=nr+1   

    wb1.save(output_path)

    #+str(nEp)+
if __name__ == "__main__":

    green = PatternFill(fill_type='solid', start_color='92D050')
    red = PatternFill(fill_type='solid', start_color='EA3916')
    n_rowFtp = 1748
    nEp = "133"
    file1 = "c:/Users/luigi.marazzi/Desktop/GG/XL/FTP/Ep133_Package_Report_20240222.xlsx"
    file2 = "c:/Users/luigi.marazzi/Desktop/GG/XL/FTP/ME_ASSET_LIST.xlsx"
    output_path = "c:/Users/luigi.marazzi/Desktop/GG/XL/FTP/FTP_Check_Ep"+str(nEp)+"_%s%s_%s%s%s.xlsx" % ( dt.hour, dt.minute, dt.day, dt.month, dt.year)

    compare_and_highlight(file1, "Sheet1", "B", file2, "SET EL PROP", "I")
